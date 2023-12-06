import openpyxl
import json
import requests
import argparse
import sys
import base64

def process_cve(finding, sheet,image_name,vulnerable_packages):
    cve_details = finding.get("cveDetails", {})
    recommendation = finding.get("remediation")
    base_score = ""
    package_id_list = cve_details.get("packageIds", [])  # Keep this as a list
    row_num = sheet.max_row + 1

    if cve_details and isinstance(cve_details, dict):
        cvss_info = cve_details.get("cvssInfo", {})
        if cvss_info and isinstance(cvss_info, dict):
            base_score = cvss_info.get("baseScore", "")

    # Extract relevant details based on package_ids
    names, versions, file_paths = [], [], []
    for pkg_id in package_id_list:
        pkg_details = vulnerable_packages.get(pkg_id, {})
        names.append(pkg_details.get("name", ""))
        versions.append(pkg_details.get("version", ""))
        file_paths.append(pkg_details.get("filePath", ""))

    sheet.cell(row=row_num, column=1, value=image_name)
    sheet.cell(row=row_num, column=2, value=finding.get("name", ''))
    sheet.cell(row=row_num, column=3, value=finding.get("description", ''))
    sheet.cell(row=row_num, column=4, value=finding.get("severity", ''))
    sheet.cell(row=row_num, column=5, value=finding.get("type", ''))
    sheet.cell(row=row_num, column=6, value=base_score)
    sheet.cell(row=row_num, column=7, value=", ".join(names))
    sheet.cell(row=row_num, column=8, value=", ".join(versions))
    sheet.cell(row=row_num, column=9, value=", ".join(file_paths))
    sheet.cell(row=row_num, column=10, value=recommendation.get("recommendation", ''))
def process_secret(finding, sheet,image_name):
    secret_details = finding.get("secretDetails", {})
    files = secret_details.get("files", [])
    recommendation = finding.get("remediation")
    for file in files:
        contents = file.get("contents", [])
        for content in contents:
            row_num = sheet.max_row + 1
            sheet.cell(row=row_num, column=1, value=image_name)
            sheet.cell(row=row_num, column=2, value=finding.get("name", ''))
            sheet.cell(row=row_num, column=3, value=finding.get("description", ''))
            sheet.cell(row=row_num, column=4, value=finding.get("severity", ''))
            sheet.cell(row=row_num, column=5, value=finding.get("type", ''))
            sheet.cell(row=row_num, column=6, value=file.get("filePath", ''))
            sheet.cell(row=row_num, column=7, value=", ".join(map(str, content.get("lines", []))))
            sheet.cell(row=row_num, column=8, value=recommendation.get("recommendation", ''))

def process_threat(finding, sheet,image_name):
    threat_details = finding.get("threatDetails", {})
    files = threat_details.get("files", [])
    recommendation = finding.get("remediation")
    for file in files:
        contents = file.get("contents", [])
        if not contents:  # if contents is empty or missing
            row_num = sheet.max_row + 1
            sheet.cell(row=row_num, column=1, value=finding.get("severity", 'None'))
            sheet.cell(row=row_num, column=2, value=image_name)
            sheet.cell(row=row_num, column=3, value=threat_details.get("type", 'None'))
            sheet.cell(row=row_num, column=4, value=threat_details.get("classification", 'None'))
            sheet.cell(row=row_num, column=5, value=finding.get("description", 'None'))
            sheet.cell(row=row_num, column=6, value=recommendation.get("recommendation", 'None'))
            sheet.cell(row=row_num, column=7, value=file.get("filePath", 'None'))
            sheet.cell(row=row_num, column=8, value='None')
            sheet.cell(row=row_num, column=9, value='None')
        else:
            for content in contents:
                row_num = sheet.max_row + 1
                sheet.cell(row=row_num, column=1, value=finding.get("severity", 'None'))
                sheet.cell(row=row_num, column=2, value=image_name)
                sheet.cell(row=row_num, column=3, value=threat_details.get("type", 'None'))
                sheet.cell(row=row_num, column=4, value=threat_details.get("classification", 'None'))
                sheet.cell(row=row_num, column=5, value=finding.get("description", 'None'))
                sheet.cell(row=row_num, column=6, value=recommendation.get("recommendation", 'None'))
                sheet.cell(row=row_num, column=7, value=file.get("filePath", 'None'))
                sheet.cell(row=row_num, column=8, value=content.get("payload", 'None'))
                sheet.cell(row=row_num, column=9, value=", ".join(map(str, content.get("lines", ['None']))))

def main():
    parser = argparse.ArgumentParser(description="Script to process CLI arguments")

    # Add the command-line arguments
    parser.add_argument("--username", required=True, help="Username for authentication")
    parser.add_argument("--password", required=True, help="Password for authentication")
    parser.add_argument("--image-lists", required=True, help="List of image names separated by commas")
    parser.add_argument("--shift-left-env-id", required=True, help="Shift left environment ID")
    args = parser.parse_args()

    # Access the parsed arguments
    username = args.username
    password = args.password
    image_list = args.image_lists.split(',')
    shift_left_env_id = args.shift_left_env_id

    print("Username:", username)
    print("Password:", password)
    print("Shift Left Environment ID:", shift_left_env_id)

    wb = openpyxl.Workbook()
    sheets = {}
    headers_for_types = {
        "CVE": ["ImageName", "name", "description", "severity", "type", "baseScore", "PackageName","PackageVersion","Image FilePath", "remediation"],
        "SECRET": ["id", "name", "description", "severity", "type", "secretFilePath", "secretLines", "remediation"],
        "THREAT": ["Severity", "name","threatType","threatClassification", "description", "remediation", "FilePath","payload","Lines"]
    }
    # Create a summary sheet to store images without results
    summary_sheet = wb.create_sheet(title="Summary")
    summary_sheet.cell(row=1, column=1, value="Image Name")
    summary_sheet.cell(row=1, column=2, value="Summary Result")
    summary_row = 2

    # Create sheets and add headers for each finding type
    for f_type, head in headers_for_types.items():
        sheets[f_type] = wb.create_sheet(title=f_type)
        for col_num, header in enumerate(head, 1):
            sheets[f_type].cell(row=1, column=col_num, value=header)

    # Loop through the image list and perform API calls
    for image_name in image_list:
        # Reset fields for each image and then append the required data
        fields = [
            {"name": "type", "value": "ShiftLeftImage"},
            {"name": "cloudAccountId", "value": shift_left_env_id},
            {"name": "name", "value": image_name}
        ]
        payload = {"filter": {"fields": fields}}
        credentials = f"{username}:{password}"
        encoded_credentials = base64.b64encode(credentials.encode()).decode()
        headers = {
            "accept": "application/json",
            "authorization": f"Basic {encoded_credentials}"
        }
        url = "https://api.eu1.dome9.com/v2/protected-asset/search"
        response = requests.post(url, json=payload, headers=headers)
        if response.status_code != 201:
            print("Failed to fetch image data. Please check connection or if the username and key are still valid.")
            sys.exit(1)

        entity_data = response.json()
        assets = entity_data.get("assets", [])
        scan_results_list = []
        found_image = False
        for asset in assets:
            if asset.get("name") == image_name:
                found_image = True
                entity_id = asset.get("entityId")
                print("Processing asset:", entity_id)
                scan_results_url = f"https://api.eu1.dome9.com/v2/vulnerability/scan-results?EntityType=Image&EntityId={entity_id}"
                scan_results_response = requests.get(scan_results_url, headers=headers)
                scan_results_data = scan_results_response.json()
                if scan_results_data.get("vulnerabilityScanFindings", []):  # Only append if there are findings
                    scan_results_list.append(scan_results_data)
                break
        if not found_image:
            summary_sheet.cell(row=summary_row, column=1, value=image_name)
            summary_sheet.cell(row=summary_row, column=2, value="Image name not found.")
            summary_row += 1
        elif not scan_results_list:
            summary_sheet.cell(row=summary_row, column=1, value=image_name)
            summary_sheet.cell(row=summary_row, column=2, value="No scan results for this image.")
            summary_row += 1
        for scan_result in scan_results_list:
            vulnerable_packages = {}
            for pkg in scan_result.get("vulnerablePackages", []):
                vulnerable_packages[pkg["id"]] = pkg

            for finding in scan_result.get("vulnerabilityScanFindings", []):
                f_type = finding['type']
                sheet = sheets.get(f_type)
                if f_type == "CVE":
                    process_cve(finding, sheet,image_name,vulnerable_packages)
                elif f_type == "THREAT":
                    process_threat(finding, sheet,image_name)
                elif f_type == "SECRET":
                    process_secret(finding, sheet,image_name)

    wb.remove(wb["Sheet"])  # Remove default sheet
    wb.save("imagereport.xlsx")

if __name__ == "__main__":
    main()